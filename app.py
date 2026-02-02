from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_cors import CORS
import requests
import json
from datetime import datetime, timedelta
import os
import re
import secrets
from functools import wraps

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


app = Flask(__name__)

# =============================================================================
# CONFIGURATION
# =============================================================================
app.config['SECRET_KEY'] = secrets.token_hex(32)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///property_tool.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)

# Initialize extensions
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
CORS(app, supports_credentials=True)

# =============================================================================
# API CONFIGURATION - Property APIs
# =============================================================================
# REGRID_API_TOKEN = "eyJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJyZWdyaWQuY29tIiwiaWF0IjoxNzY4ODAxODAyLCJleHAiOjE3NzEzOTM4MDIsInUiOjY0MDQyMCwiZyI6MjMxNTMsImNhcCI6InBhOnRzOnBzOmJmOm1hOnR5OmVvOnpvOnNiIn0.ccYCDK63ya2biBGqejR2itkO8IOZjxs77aGIkXY7poM"

REGRID_API_TOKEN = "eyJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJyZWdyaWQuY29tIiwiaWF0IjoxNzcwMDA3MDE3LCJleHAiOjE3NzI1OTkwMTcsInUiOjY2MTk1OCwiZyI6MjMxNTMsImNhcCI6InBhOnRzOnBzOmJmOm1hOnR5OmVvOnpvOnNiIn0.-mp5f-cYJjwuv4VhWdSCXXZg2fbNz_L9ch1DC-eNFas"

SMARTY_AUTH_ID = "8b477fb8-a789-4613-99d1-007a5820ee83"
SMARTY_AUTH_TOKEN = "HpReViw8xnxls7QppqnM"
MELISSA_LICENSE_KEY = "FuKFleH4vbD8kX7LbUW7WE**"

REGRID_BASE_URL = "https://app.regrid.com/api/v2/parcels/address"
SMARTY_BASE_URL = "https://us-enrichment.api.smarty.com/lookup/search/property/principal"
MELISSA_BASE_URL = "https://property.melissadata.net/v4/WEB/LookupProperty"

# Create directories
SAVE_DIRECTORY = 'saved_properties'
if not os.path.exists(SAVE_DIRECTORY):
    os.makedirs(SAVE_DIRECTORY)

# =============================================================================
# DATABASE MODELS
# =============================================================================
class User(db.Model):
    """User model for authentication"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    
    def __repr__(self):
        return f'<User {self.email}>'


class PasswordReset(db.Model):
    """Password reset tokens"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    token = db.Column(db.String(255), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, default=False)
    
    user = db.relationship('User', backref='reset_tokens')



class APIStatisticsHistory(db.Model):
    """Store API statistics for each address lookup"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    address = db.Column(db.String(500), nullable=False)
    
    # Regrid stats
    regrid_available = db.Column(db.Integer, default=0)
    regrid_unavailable = db.Column(db.Integer, default=0)
    regrid_total = db.Column(db.Integer, default=0)
    
    # Smarty stats
    smarty_available = db.Column(db.Integer, default=0)
    smarty_unavailable = db.Column(db.Integer, default=0)
    smarty_total = db.Column(db.Integer, default=0)
    
    # Melissa stats
    melissa_available = db.Column(db.Integer, default=0)
    melissa_unavailable = db.Column(db.Integer, default=0)
    melissa_total = db.Column(db.Integer, default=0)
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='api_statistics')
    
    def to_dict(self):
        return {
            'id': self.id,
            'address': self.address,
            'regrid': {
                'available': self.regrid_available,
                'unavailable': self.regrid_unavailable,
                'total': self.regrid_total
            },
            'smarty': {
                'available': self.smarty_available,
                'unavailable': self.smarty_unavailable,
                'total': self.smarty_total
            },
            'melissa': {
                'available': self.melissa_available,
                'unavailable': self.melissa_unavailable,
                'total': self.melissa_total
            },
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S')
        }


# =============================================================================
# ADDRESS VALIDATION - BEFORE API CALLS
# =============================================================================


def normalize_address_format(address):
    """
    Normalizes multiple formats into:
    Street, City, STATE ZIP

    Supports:
    - 519 Raton Pass, Irving, TX 75063
    - 519 Raton Pass Irving TX 75063
    - Irving 519 Raton Pass TX 75063
    - 519 Raton Pass Irving TX 75063 USA
    - 436 Red River Trl #1112	Irving	75063	TX (tab/space separated, ZIP before STATE)
    """
    address = address.strip()
    # Replace tabs with spaces
    address = re.sub(r'\t+', ' ', address)
    # Replace multiple spaces with single space
    address = re.sub(r'\s+', ' ', address)

    # Remove country if present at end
    address = re.sub(r'\b(US|USA|UAS|UNITED STATES)\b$', '', address, flags=re.IGNORECASE).strip()

    # If already formatted with commas
    if ',' in address:
        return address

    parts = address.split()

    if len(parts) < 4:
        return address

    # Try to identify components
    state = None
    zip_code = None
    
    # Look for ZIP (5 digits or 5+4 digits)
    for i, part in enumerate(parts):
        if re.match(r'^\d{5}(-\d{4})?$', part):
            zip_code = part.split('-')[0]
            zip_index = i
            break
    
    if not zip_code:
        return address
    
    # Look for STATE (2 letters)
    # Check both before and after ZIP
    state_candidates = []
    for i, part in enumerate(parts):
        if re.match(r'^[A-Za-z]{2}$', part):
            state_candidates.append((i, part.upper()))
    
    if not state_candidates:
        return address
    
    # Pick the state closest to ZIP
    state_index, state = min(state_candidates, key=lambda x: abs(x[0] - zip_index))
    
    # Remove ZIP and STATE from parts
    remaining_parts = [p for i, p in enumerate(parts) if i not in [zip_index, state_index]]
    
    if len(remaining_parts) < 2:
        return address
    
    # Check if first part starts with digit (street address)
    first_part = remaining_parts[0]
    
    if first_part[0].isdigit():
        # Format: Street City
        # City is last part
        city = remaining_parts[-1]
        street = ' '.join(remaining_parts[:-1])
    elif first_part[0].isalpha():
        # Format: City Street
        city = remaining_parts[0]
        street = ' '.join(remaining_parts[1:])
    else:
        return address

    return f"{street}, {city}, {state} {zip_code}"


def validate_address_format(address):
    original_address = address
    address = normalize_address_format(address).strip()

    print(f"[VALIDATION] Original: '{original_address}'")
    print(f"[VALIDATION] Normalized: '{address}'")

    if len(address) < 10:
        return False, "Address too short (minimum 10 characters)"

    if ',' not in address:
        return False, "Address format incomplete (need: Street, City, State ZIP)"

    parts = [p.strip() for p in address.split(',') if p.strip()]

    if len(parts) < 2:
        return False, "Address must include at least Street and City"

    street_part = parts[0]

    if not re.search(r'\d', street_part):
        return False, "Street address must include a house number"

    house_match = re.match(r'^(\d+)', street_part)
    if house_match:
        house_num = int(house_match.group(1))
        if house_num < 1 or house_num > 99999:
            return False, "Invalid house number"
        if re.match(r'^(\d)\1+$', house_match.group(1)):
            return False, "Invalid house number pattern"

    if len(street_part.split()) < 2:
        # Allow single word if it contains unit number like "#1112"
        if not re.search(r'#\d+', street_part):
            return False, "Street name missing"

    if len(parts) >= 2:
        # Check last part for state/zip
        last = parts[-1]

        if not re.search(r'\b[A-Z]{2}\b|\b\d{5}(?:-\d{4})?\b', last, re.IGNORECASE):
            return False, "State or ZIP missing"

        zip_match = re.search(r'\b(\d{5})(?:-\d{4})?\b', last)
        if zip_match:
            if re.match(r'^(\d)\1{4}$', zip_match.group(1)):
                return False, "Invalid ZIP code"

    invalid_patterns = [
        (r'test\s*address', 'Test addresses not allowed'),
        (r'example', 'Example addresses not allowed'),
        (r'fake', 'Fake addresses not allowed'),
        (r'asdf|qwerty', 'Invalid address format'),
        (r'^[\d\s,]+$', 'Address must include text'),
    ]

    for pattern, error_msg in invalid_patterns:
        if re.search(pattern, address, re.IGNORECASE):
            return False, error_msg

    return True, None

# =============================================================================
# AUTHENTICATION HELPER FUNCTIONS
# =============================================================================
def validate_email(email):
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_password(password):
    """Validate password strength"""
    if len(password) < 8:
        return False, "Password must be at least 8 characters long"
    if not re.search(r'[A-Z]', password):
        return False, "Password must contain at least one uppercase letter"
    if not re.search(r'[a-z]', password):
        return False, "Password must contain at least one lowercase letter"
    if not re.search(r'\d', password):
        return False, "Password must contain at least one number"
    return True, "Password is valid"


def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function


# =============================================================================
# PROPERTY API HELPER FUNCTIONS
# =============================================================================
def sanitize_filename(address):
    """Convert address to safe filename WITHOUT timestamp"""
    safe_name = re.sub(r'[^\w\s-]', '', address)
    safe_name = re.sub(r'[-\s]+', '_', safe_name)
    safe_name = safe_name[:150]
    return f"{safe_name}.json"

FIELD_MAPPING = {
    # === BASIC PROPERTY INFO ===
    'year_built': ['yearbuilt', 'year_built', 'PropertyUseInfo_YearBuilt'],
    'bedrooms': ['num_bedrooms', 'bedrooms', 'IntRoomInfo_BedroomsCount'],
    'bathrooms_total': ['num_bath', 'bathrooms_total', 'IntRoomInfo_BathCount'],
    'bathrooms_partial': ['num_bath_partial', 'bathrooms_partial', None],
    'rooms_total': ['numrooms', None, 'IntRoomInfo_RoomsCount'],
    'building_sqft': ['area_building', 'building_sqft', 'PropertySize_AreaBuilding'],
    'living_area_sqft': [None, 'gross_sqft', None],
    'stories': ['numstories', 'stories_number', None],
    'units': ['numunits', None, None],
    'structure_style': ['structstyle', None, None],
    
    # === LOT & LAND ===
    'lot_sqft': ['ll_gissqft', 'lot_sqft', 'PropertySize_AreaLotSF'],
    'lot_acres': ['ll_gisacre', 'acres', 'PropertySize_AreaLotAcres'],
    'deeded_acres': ['deeded_acres', None, None],
    'building_footprint_sqft': ['ll_bldg_footprint_sqft', None, None],
    'building_count': ['ll_bldg_count', None, None],
    
    # === VALUATION & ASSESSMENT ===
    'assessed_value_total': ['parval', 'assessed_value', 'Tax_AssessedValueTotal'],
    'assessed_land_value': ['landval', 'assessed_land_value', None],
    'assessed_improvement_value': ['improvval', 'assessed_improvement_value', None],
    'market_value_total': [None, 'total_market_value', 'Tax_MarketValueTotal'],
    'market_land_value': [None, 'market_land_value', None],
    'market_improvement_value': [None, 'market_improvement_value', None],
    'assessed_improvement_percent': [None, 'assessed_improvement_percent', None],
    'market_improvement_percent': [None, 'market_improvement_percent', None],
    'agricultural_value': ['agval', None, None],
    'parcel_value_type': ['parvaltype', None, None],
    
    # === TAX INFO ===
    'tax_amount': ['taxamt', 'tax_billed_amount', 'Tax_TaxBilledAmount'],
    'tax_year': ['taxyear', 'tax_fiscal_year', 'Tax_TaxFiscalYear'],
    'tax_assess_year': [None, 'tax_assess_year', None],
    'tax_jurisdiction': [None, 'tax_jurisdiction', None],
    'tax_rate_area': [None, 'tax_rate_area', None],
    'homestead_exemption': ['homestead_exemption', 'homeowner_tax_exemption', None],
    
    # === SALES INFO ===
    'sale_price': ['saleprice', None, 'SaleInfo_DeedLastSalePrice'],
    'sale_date': ['saledate', 'sale_date', 'SaleInfo_DeedLastSaleDate'],
    'prior_sale_date': [None, 'prior_sale_date', 'SaleInfo_AssessorPriorSaleDate'],
    'prior_sale_price': [None, None, 'SaleInfo_AssessorPriorSaleAmount'],
    'last_ownership_transfer_date': ['last_ownership_transfer_date', 'ownership_transfer_date', None],
    'deed_sale_date': [None, 'deed_sale_date', None],
    'assessor_last_sale_date': [None, None, 'SaleInfo_AssessorLastSaleDate'],
    'assessor_last_sale_amount': [None, None, 'SaleInfo_AssessorLastSaleAmount'],
    
    # === OWNER INFO ===
    'owner_name': ['owner', 'owner_full_name', 'PrimaryOwner_Name1Full'],
    'owner_first_name': ['ownfrst', 'first_name', None],
    'owner_last_name': ['ownlast', 'last_name', None],
    'owner_name_2': ['owner2', 'owner_full_name_2', 'SecondaryOwner_Name1Full'],
    'owner_first_name_2': [None, 'first_name_2', None],
    'owner_last_name_2': [None, 'last_name_2', None],
    'owner_name_3': ['owner3', 'owner_full_name_3', None],
    'owner_name_4': ['owner4', None, None],
    'owner_type': ['owntype', 'ownership_type', None],
    'owner_occupancy_status': [None, 'owner_occupancy_status', None],
    'company_flag': [None, 'company_flag', None],
    'trust_description': [None, 'trust_description', None],
    'previous_owner': ['previous_owner', None, None],
    'unmodified_owner': ['unmodified_owner', None, None],
    
    # === DEED OWNER INFO ===
    'deed_owner_full_name': [None, 'deed_owner_full_name', None],
    'deed_owner_first_name': [None, 'deed_owner_first_name', None],
    'deed_owner_last_name': [None, 'deed_owner_last_name', None],
    'deed_owner_full_name_2': [None, 'deed_owner_full_name2', None],
    'deed_owner_full_name_3': [None, 'deed_owner_full_name3', None],
    'deed_document_number': [None, 'deed_document_number', None],
    'deed_transaction_id': [None, 'deed_transaction_id', None],
    
    # === MAILING ADDRESS ===
    'mail_address': ['mailadd', 'contact_full_address', 'OwnerAddress_Address1'],
    'mail_address_2': ['mail_address2', None, None],
    'mail_city': ['mail_city', 'contact_city', None],
    'mail_state': ['mail_state2', 'contact_state', None],
    'mail_zip': ['mail_zip', 'contact_zip', None],
    'mail_zip4': [None, 'contact_zip4', None],
    'mail_country': ['mail_country', None, None],
    'mail_care_of': ['careof', None, None],
    'mail_house_number': ['mail_addno', 'contact_house_number', None],
    'mail_street_name': ['mail_addstr', 'contact_street_name', None],
    'mail_street_suffix': ['mail_addsttyp', 'contact_suffix', None],
    'mail_carrier_route': [None, 'contact_crrt', None],
    
    # === PROPERTY ADDRESS ===
    'property_address': ['address', 'property_address_full', None],
    'property_house_number': ['saddno', 'property_address_house_number', None],
    'property_street_name': ['saddstr', 'property_address_street_name', None],
    'property_street_suffix': ['saddsttyp', 'property_address_street_suffix', None],
    'property_unit': ['sunit', None, None],
    'property_city': ['scity', 'property_address_city', None],
    'property_state': ['state2', 'property_address_state', None],
    'property_zip': ['szip', 'property_address_zipcode', None],
    'property_zip4': [None, 'property_address_zip_4', None],
    'property_carrier_route': [None, 'property_address_carrier_route_code', None],
    'address_source': ['address_source', None, None],
    'situs_county': [None, 'situs_county', None],
    'situs_state': [None, 'situs_state', None],
    
    # === PARCEL INFO ===
    'parcel_number': ['parcelnumb', 'parcel_raw_number', 'Parcel_UnformattedAPN'],
    'parcel_number_no_formatting': ['parcelnumb_no_formatting', None, None],
    'parcel_number_alternate': [None, 'parcel_number_alternate', None],
    'account_number': ['account_number', None, None],
    'state_parcel_number': ['state_parcelnumb', None, None],
    'tax_id': ['tax_id', None, None],
    'parcel_number_year_added': [None, 'parcel_number_year_added', None],
    'parcel_number_year_change': [None, 'parcel_number_year_change', None],
    'formatted_apn': [None, None, 'Parcel_FormattedAPN'],
    
    # === LOCATION & GEOGRAPHY ===
    'latitude': ['lat', 'latitude', None],
    'longitude': ['lon', 'longitude', None],
    'fips_code': ['geoid', 'fips_code', 'Parcel_FIPSCode'],
    'county': ['county', None, None],
    'city': ['city', None, None],
    'census_tract': ['census_tract', 'census_tract', None],
    'census_block': ['census_block', 'census_block', None],
    'census_blockgroup': ['census_blockgroup', 'census_block_group', None],
    'census_zcta': ['census_zcta', None, None],
    'elevation_feet': [None, 'elevation_feet', None],
    'highest_parcel_elevation': ['highest_parcel_elevation', None, None],
    'lowest_parcel_elevation': ['lowest_parcel_elevation', None, None],
    'roughness_rating': ['roughness_rating', None, None],
    
    # === LAND USE & ZONING ===
    'use_code': ['usecode', 'land_use_code', None],
    'use_description': ['usedesc', 'land_use_standard', None],
    'land_use_group': [None, 'land_use_group', None],
    'zoning': ['zoning', 'zoning', None],
    'zoning_description': ['zoning_description', 'zoning_description', None],
    'zoning_type': ['zoning_type', None, None],
    'zoning_subtype': ['zoning_subtype', None, None],
    
    # === LEGAL DESCRIPTION ===
    'legal_description': ['legaldesc', 'legal_description', None],
    'subdivision': ['subdivision', 'subdivision', None],
    'neighborhood': ['neighborhood', None, None],
    'neighborhood_code': ['neighborhood_code', 'neighborhood_code', None],
    'plat': ['plat', None, None],
    'block': ['block', 'block1', None],
    'lot': ['lot', 'lot_1', None],
    'book': ['book', None, None],
    'page': ['page', None, None],
    'tract_number': [None, 'tract_number', None],
    'depth_linear_footage': [None, 'depth_linear_footage', None],
    'width_linear_footage': [None, 'width_linear_footage', None],
    
    # === BUILDING FEATURES ===
    'construction_type': [None, 'construction_type', None],
    'building_definition_code': [None, 'building_definition_code', None],
    'exterior_walls': [None, 'exterior_walls', None],
    'roof_cover': [None, 'roof_cover', None],
    'roof_frame': [None, 'roof_frame', None],
    'foundation': [None, 'foundation', None],
    'flooring': [None, 'flooring', None],
    
    # === AMENITIES & FEATURES ===
    'garage': [None, 'garage', None],
    'garage_sqft': ['garage_sqft', 'garage_sqft', None],
    'parking_spaces': [None, 'parking_spaces', None],
    'pool': [None, 'pool', None],
    'deck': [None, 'deck', None],
    'fence': [None, 'fence', None],
    'fireplace': [None, 'fireplace', None],
    'fireplace_number': [None, 'fireplace_number', None],
    'air_conditioner': [None, 'air_conditioner', None],
    'heat': [None, 'heat', None],
    'sprinklers': [None, 'sprinklers', None],
    'fire_sprinklers': [None, 'fire_sprinklers_flag', None],
    'wet_bar': [None, 'wet_bar', None],
    
    # === CENSUS & SCHOOL DISTRICTS ===
    'census_elementary_school_district': ['census_elementary_school_district', None, None],
    'census_secondary_school_district': ['census_secondary_school_district', None, None],
    'census_unified_school_district': ['census_unified_school_district', None, None],
    
    # === FLOOD & RISK ===
    'fema_flood_zone': ['fema_flood_zone', None, None],
    'fema_flood_zone_subtype': ['fema_flood_zone_subtype', None, None],
    'fema_nri_risk_rating': ['fema_nri_risk_rating', None, None],
    
    # === DEMOGRAPHICS (Regrid Premium) ===
    'population_density': ['population_density', None, None],
    'median_household_income': ['median_household_income', None, None],
    'housing_affordability_index': ['housing_affordability_index', None, None],
    
    # === USPS VALIDATION ===
    'dpv_status': ['dpv_status', None, None],
    'dpv_codes': ['dpv_codes', None, None],
    'dpv_type': ['dpv_type', None, None],
    'usps_vacancy': ['usps_vacancy', None, None],
    'rdi': ['rdi', None, None],
    
    # === THIRD PARTY IDS ===
    'precisely_id': ['precisely_id', None, None],
    'placekey': ['placekey', None, None],
    'smarty_key': [None, 'smarty_key', None],
    'll_uuid': ['ll_uuid', None, None],
    
    # === METADATA ===
    'll_last_refresh': ['ll_last_refresh', None, None],
    'publication_date': [None, 'publication_date', None],
    'recording_date': [None, 'recording_date', None],
    'instrument_date': [None, 'instrument_date', None],
    'assessor_last_update': [None, 'assessor_last_update', None],
    'assessor_taxroll_update': [None, 'assessor_taxroll_update', None],
    
    # === MORTGAGE & DEED ===
    'mortgage_amount': [None, None, 'CurrentDeed_MortgageAmount'],
    'mortgage_date': [None, None, 'CurrentDeed_MortgageDate'],
    'lender_name': [None, None, 'CurrentDeed_LenderName'],
    'document_type_description': [None, 'document_type_description', None],
    
    # === CBSA & MSA ===
    'cbsa_code': [None, 'cbsa_code', None],
    'cbsa_name': [None, 'cbsa_name', None],
    'msa_code': [None, 'msa_code', None],
    'msa_name': [None, 'msa_name', None],
    'metro_division': [None, 'metro_division', None],
    'combined_statistical_area': [None, 'combined_statistical_area', None],
    
    # === QOZ (Qualified Opportunity Zone) ===
    'qoz': ['qoz', None, None],
    'qoz_tract': ['qoz_tract', None, None],
    
    # === MATCH & SOURCE INFO ===
    'match_type': [None, 'match_type', None],
    'data_set_name': [None, 'data_set_name', None],
    'data_subset_name': [None, 'data_subset_name', None],
}

def parse_address(address_string):
    """Parse address"""
    parts = [p.strip() for p in address_string.split(',')]
    parsed = {
        'street': parts[0] if len(parts) > 0 else '',
        'city': parts[1] if len(parts) > 1 else '',
        'state': parts[2] if len(parts) > 2 else '',
        'zipcode': parts[3] if len(parts) > 3 else ''
    }
    return parsed


def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            if len(v) > 0 and isinstance(v[0], dict):
                for i, item in enumerate(v):
                    items.extend(flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
            else:
                items.append((new_key, str(v)))
        else:
            items.append((new_key, v))
    return dict(items)


def fetch_regrid_data(address):
    """Fetch from Regrid API"""
    try:
        headers = {"accept": "application/json", "x-regrid-token": REGRID_API_TOKEN}
        params = {"query": address, "limit": 1}
        response = requests.get(REGRID_BASE_URL, headers=headers, params=params, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if "parcels" in data:
                parcels = data["parcels"]
                if isinstance(parcels, dict) and "features" in parcels:
                    if len(parcels["features"]) > 0:
                        props = parcels["features"][0]["properties"]
                        if "fields" in props:
                            return flatten_dict(props["fields"])
        return {}
    except Exception as e:
        print(f"[Regrid] Error: {str(e)}")
        return {}


def fetch_smarty_data(address):
    """Fetch from Smarty API"""
    try:
        parsed = parse_address(address)
        params = {
            'auth-id': SMARTY_AUTH_ID,
            'auth-token': SMARTY_AUTH_TOKEN,
            'street': parsed['street'],
            'city': parsed['city'],
            'state': parsed['state']
        }
        if parsed['zipcode']:
            params['zipcode'] = parsed['zipcode']
        response = requests.get(SMARTY_BASE_URL, params=params, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, list) and len(data) > 0:
                result = data[0]
                flattened = {}
                if 'attributes' in result:
                    flattened.update(result['attributes'])
                return flattened
        return {}
    except Exception as e:
        print(f"[Smarty] Error: {str(e)}")
        return {}


def fetch_melissa_data(address):
    """Fetch from Melissa API"""
    try:
        params = {'id': MELISSA_LICENSE_KEY, 'ff': address, 'format': 'json'}
        response = requests.get(MELISSA_BASE_URL, params=params, timeout=30)
        if response.status_code == 200:
            data = response.json()
            if 'Records' in data and len(data['Records']) > 0:
                return flatten_dict(data['Records'][0])
        return {}
    except Exception as e:
        print(f"[Melissa] Error: {str(e)}")
        return {}


def is_valid_address_data(regrid_data, smarty_data, melissa_data):
    """Check if address returned valid data"""
    if not regrid_data and not smarty_data and not melissa_data:
        return False, "Invalid address - No data found from any API"
    has_valid_data = False
    if regrid_data and any(key in regrid_data for key in ['address', 'parcelnumb', 'owner', 'yearbuilt']):
        has_valid_data = True
    if smarty_data and any(key in smarty_data for key in ['property_address_full', 'parcel_raw_number', 'year_built']):
        has_valid_data = True
    if melissa_data and any(key in melissa_data for key in ['PropertyUseInfo_YearBuilt', 'Parcel_UnformattedAPN']):
        has_valid_data = True
    if not has_valid_data:
        return False, "Invalid address - No valid property data found"
    return True, "Valid address"


def calculate_api_statistics(regrid_data, smarty_data, melissa_data, comparison):
    """
    Calculate statistics for each API - ALWAYS 185 FIELDS TOTAL
    This function counts directly from raw API data, not from the comparison array
    """
    # Initialize stats - TOTAL is always 185 (number of fields in FIELD_MAPPING)
    TOTAL_FIELDS = len(FIELD_MAPPING)  # Dynamically get count
    stats = {
        'regrid': {'available': 0, 'unavailable': 0, 'total': TOTAL_FIELDS},
        'smarty': {'available': 0, 'unavailable': 0, 'total': TOTAL_FIELDS},
        'melissa': {'available': 0, 'unavailable': 0, 'total': TOTAL_FIELDS}
    }
    
    # Count through EVERY field in FIELD_MAPPING (all 217 fields)
    for common_name, api_names in FIELD_MAPPING.items():
        regrid_name, smarty_name, melissa_name = api_names
        
        # Check Regrid - field must exist in mapping AND have data
        if regrid_name and regrid_name in regrid_data and regrid_data.get(regrid_name):
            stats['regrid']['available'] += 1
        else:
            stats['regrid']['unavailable'] += 1
        
        # Check Smarty - field must exist in mapping AND have data
        if smarty_name and smarty_name in smarty_data and smarty_data.get(smarty_name):
            stats['smarty']['available'] += 1
        else:
            stats['smarty']['unavailable'] += 1
        
        # Check Melissa - field must exist in mapping AND have data
        if melissa_name and melissa_name in melissa_data and melissa_data.get(melissa_name):
            stats['melissa']['available'] += 1
        else:
            stats['melissa']['unavailable'] += 1
    
    # Debug output to verify counts
    print(f"[STATS] Regrid: {stats['regrid']['available']} + {stats['regrid']['unavailable']} = {stats['regrid']['total']}")
    print(f"[STATS] Smarty: {stats['smarty']['available']} + {stats['smarty']['unavailable']} = {stats['smarty']['total']}")
    print(f"[STATS] Melissa: {stats['melissa']['available']} + {stats['melissa']['unavailable']} = {stats['melissa']['total']}")
    
    return stats


def analyze_field_discrepancies(comparison):
    """
    Analyze field discrepancies across APIs
    Returns ONLY fields where values are different
    Sorted by status priority, then by field name
    """
    discrepancies = []
    
    for row in comparison:
        field_name = row['field_name']
        regrid = row['regrid'].strip() if row['regrid'] else ""
        smarty = row['smarty'].strip() if row['smarty'] else ""
        melissa = row['melissa'].strip() if row['melissa'] else ""
        
        # Count how many APIs have data
        apis_with_data = sum([bool(regrid), bool(smarty), bool(melissa)])
        
        # Skip if less than 2 APIs have data (can't compare)
        if apis_with_data < 2:
            continue
        
        # Normalize values for comparison (case-insensitive, strip spaces)
        regrid_norm = regrid.lower().strip() if regrid else None
        smarty_norm = smarty.lower().strip() if smarty else None
        melissa_norm = melissa.lower().strip() if melissa else None
        
        # Determine if values are different
        is_different = False
        status = ""
        
        if apis_with_data == 2:
            # Two APIs have data - check if they differ
            values = [v for v in [regrid_norm, smarty_norm, melissa_norm] if v is not None]
            if values[0] != values[1]:
                is_different = True
                status = "Values Differ"
        
        elif apis_with_data == 3:
            # All three APIs have data
            if regrid_norm == smarty_norm == melissa_norm:
                # All match - skip this field
                is_different = False
            else:
                is_different = True
                # Check how many are different
                if regrid_norm != smarty_norm and regrid_norm != melissa_norm and smarty_norm != melissa_norm:
                    status = "All Different"
                else:
                    # 2 match, 1 differs
                    if regrid_norm == smarty_norm:
                        status = "Melissa Differs"
                    elif regrid_norm == melissa_norm:
                        status = "Smarty Differs"
                    elif smarty_norm == melissa_norm:
                        status = "Regrid Differs"
        
        # Only add fields where values are different
        if is_different:
            discrepancies.append({
                'field_name': field_name,
                'status': status,
                'regrid': regrid,
                'smarty': smarty,
                'melissa': melissa,
                'apis_with_data': apis_with_data
            })
    
    # Define status priority for sorting
    status_priority = {
        "All Different": 1,
        "Values Differ": 2,
        "Smarty Differs": 3,
        "Regrid Differs": 4,
        "Melissa Differs": 5
    }
    
    # Sort by status priority first, then by field name
    discrepancies.sort(key=lambda x: (status_priority.get(x['status'], 999), x['field_name']))
    
    return discrepancies


# =============================================================================
# ROUTES - AUTHENTICATION PAGES
# =============================================================================
@app.route('/login')
def login_page():
    """Login page"""
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/register')
def register_page():
    """Register page"""
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')


# =============================================================================
# ROUTES - PROPERTY TOOL (PROTECTED)
# =============================================================================
@app.route('/')
@login_required
def index():
    """Main property tool page (protected)"""
    user = User.query.get(session['user_id'])
    return render_template('index.html', user=user)


# =============================================================================
# API ROUTES - AUTHENTICATION
# =============================================================================
@app.route('/api/register', methods=['POST'])
def register():
    """Register new user"""
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        confirm_password = data.get('confirm_password', '')
        
        if not all([name, email, password, confirm_password]):
            return jsonify({'success': False, 'error': 'All fields are required'}), 400
        
        if not validate_email(email):
            return jsonify({'success': False, 'error': 'Invalid email format'}), 400
        
        if User.query.filter_by(email=email).first():
            return jsonify({'success': False, 'error': 'Email already registered'}), 400
        
        is_valid, message = validate_password(password)
        if not is_valid:
            return jsonify({'success': False, 'error': message}), 400
        
        if password != confirm_password:
            return jsonify({'success': False, 'error': 'Passwords do not match'}), 400
        
        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        new_user = User(name=name, email=email, password_hash=password_hash)
        db.session.add(new_user)
        db.session.commit()
        
        print(f"[REGISTER] New user: {email}")
        return jsonify({'success': True, 'message': 'Account created! Please log in.'}), 201
    except Exception as e:
        db.session.rollback()
        print(f"[REGISTER ERROR] {str(e)}")
        return jsonify({'success': False, 'error': 'Registration failed'}), 500


@app.route('/api/login', methods=['POST'])
def login():
    """Login user"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        remember_me = data.get('remember_me', False)
        
        if not all([email, password]):
            return jsonify({'success': False, 'error': 'Email and password required'}), 400
        
        user = User.query.filter_by(email=email).first()
        if not user or not bcrypt.check_password_hash(user.password_hash, password):
            return jsonify({'success': False, 'error': 'Invalid email or password'}), 401
        
        if not user.is_active:
            return jsonify({'success': False, 'error': 'Account deactivated'}), 401
        
        user.last_login = datetime.utcnow()
        db.session.commit()
        
        session.permanent = remember_me
        session['user_id'] = user.id
        session['user_email'] = user.email
        session['user_name'] = user.name
        
        print(f"[LOGIN] {email}")
        return jsonify({'success': True, 'message': 'Login successful!'}), 200
    except Exception as e:
        print(f"[LOGIN ERROR] {str(e)}")
        return jsonify({'success': False, 'error': 'Login failed'}), 500


@app.route('/api/logout', methods=['POST'])
@login_required
def logout():
    """Logout user"""
    try:
        session.clear()
        return jsonify({'success': True, 'message': 'Logged out'}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': 'Logout failed'}), 500


@app.route('/api/check-session', methods=['GET'])
def check_session():
    """Check if user is logged in"""
    if 'user_id' in session:
        return jsonify({
            'success': True,
            'logged_in': True,
            'user': {
                'id': session['user_id'],
                'name': session.get('user_name'),
                'email': session.get('user_email')
            }
        }), 200
    return jsonify({'success': True, 'logged_in': False}), 200


# =============================================================================
# API ROUTES - PROPERTY TOOL (PROTECTED)
# =============================================================================

@app.route('/api/fetch-data', methods=['POST'])
@login_required
def fetch_data():
    """Fetch property data (protected) - WITH ADDRESS VALIDATION FIRST"""
    try:
        data = request.get_json()
        address = data.get('address', '')
        
        if not address:
            return jsonify({"success": False, "error": "Address required"}), 400
        
        # ===== VALIDATE ADDRESS FORMAT BEFORE API CALLS =====
        is_valid, validation_error = validate_address_format(address)
        if not is_valid:
            print(f"[VALIDATION REJECTED] {address} - {validation_error}")
            return jsonify({
                "success": False,
                "error": f"Invalid Address: {validation_error}",
                "is_invalid_address": True
            }), 400
        
        print(f"[VALIDATION PASSED] {address} - Proceeding to API calls")
        # ===== END VALIDATION - NOW CALL APIs =====
        
        regrid_data = fetch_regrid_data(address)
        smarty_data = fetch_smarty_data(address)
        melissa_data = fetch_melissa_data(address)
        
        is_valid_data, validation_message = is_valid_address_data(regrid_data, smarty_data, melissa_data)
        if not is_valid_data:
            return jsonify({
                "success": False,
                "error": validation_message,
                "is_invalid_address": True
            }), 400
        
        # ===== ALWAYS SHOW ALL 217 FIELDS =====
        comparison = []
        for common_name, api_names in FIELD_MAPPING.items():
            regrid_name, smarty_name, melissa_name = api_names
            regrid_val = str(regrid_data.get(regrid_name, "")) if regrid_name and regrid_data.get(regrid_name) else ""
            smarty_val = str(smarty_data.get(smarty_name, "")) if smarty_name and smarty_data.get(smarty_name) else ""
            melissa_val = str(melissa_data.get(melissa_name, "")) if melissa_name and melissa_data.get(melissa_name) else ""
            
            # Always add ALL 217 fields to comparison
            comparison.append({
                "field_name": common_name,
                "regrid": regrid_val,
                "smarty": smarty_val,
                "melissa": melissa_val
            })
        
        # Calculate statistics from raw API data
        api_stats = calculate_api_statistics(regrid_data, smarty_data, melissa_data, comparison)
        discrepancies = analyze_field_discrepancies(comparison)
        
        try:
            api_stat_record = APIStatisticsHistory(
                user_id=session['user_id'],
                address=address,
                regrid_available=api_stats['regrid']['available'],
                regrid_unavailable=api_stats['regrid']['unavailable'],
                regrid_total=api_stats['regrid']['total'],
                smarty_available=api_stats['smarty']['available'],
                smarty_unavailable=api_stats['smarty']['unavailable'],
                smarty_total=api_stats['smarty']['total'],
                melissa_available=api_stats['melissa']['available'],
                melissa_unavailable=api_stats['melissa']['unavailable'],
                melissa_total=api_stats['melissa']['total']
            )
            db.session.add(api_stat_record)
            db.session.commit()
            print(f"[API STATS] Saved for address: {address}")
        except Exception as stat_error:
            print(f"[API STATS ERROR] {str(stat_error)}")
            db.session.rollback()

        
        return jsonify({
            "success": True,
            "address": address,
            "data": comparison,
            "api_statistics": api_stats,
            "discrepancies": discrepancies
        })
    except Exception as e:
        print(f"[ERROR] {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/save-selection', methods=['POST'])
@login_required
def save_selection():
    """Save field selections (protected)"""
    try:
        data = request.get_json()
        selections = data.get('selections', {})
        address = data.get('address', '')
        
        if not address or not selections:
            return jsonify({"success": False, "error": "Invalid data"}), 400
        
        merged_data = {field: data['value'] for field, data in selections.items()}
        output = {
            "address": address,
            "saved_at": datetime.now().isoformat(),
            "user_id": session['user_id'],
            "user_email": session['user_email'],
            "selections": selections,
            "merged_data": merged_data
        }
        
        filename = sanitize_filename(address)
        filepath = os.path.join(SAVE_DIRECTORY, filename)
        
        with open(filepath, 'w') as f:
            json.dump(output, f, indent=2)
        
        print(f"[SAVE] {len(selections)} fields saved by {session['user_email']}")
        return jsonify({"success": True, "message": f"Saved to {filename}", "count": len(selections)}), 200
    except Exception as e:
        print(f"[SAVE ERROR] {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500
    


@app.route('/api/get-statistics-history', methods=['GET'])
@login_required
def get_statistics_history():
    """Get all API statistics history for current user"""
    try:
        user_id = session['user_id']
        records = APIStatisticsHistory.query.filter_by(user_id=user_id).order_by(
            APIStatisticsHistory.created_at.desc()
        ).all()
        
        history = [record.to_dict() for record in records]
        
        return jsonify({
            "success": True,
            "history": history,
            "total_records": len(history)
        }), 200
    except Exception as e:
        print(f"[STATS HISTORY ERROR] {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/delete-statistics', methods=['DELETE'])
@login_required
def delete_statistics():
    """Delete a specific API statistics history record"""
    try:
        data = request.get_json()
        record_id = data.get('record_id')
        user_id = session['user_id']
        
        if not record_id:
            return jsonify({"success": False, "error": "Record ID required"}), 400
        
        # Find and verify ownership
        record = APIStatisticsHistory.query.filter_by(
            id=record_id,
            user_id=user_id
        ).first()
        
        if not record:
            return jsonify({"success": False, "error": "Record not found or access denied"}), 404
        
        # Delete the record
        db.session.delete(record)
        db.session.commit()
        
        print(f"[DELETE STATS] Record {record_id} deleted by user {user_id}")
        return jsonify({
            "success": True,
            "message": f"Deleted record for {record.address}"
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"[DELETE STATS ERROR] {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export-statistics-excel', methods=['GET'])
@login_required
def export_statistics_excel():
    """Export API statistics history to Excel"""
    try:
        user_id = session['user_id']
        records = APIStatisticsHistory.query.filter_by(user_id=user_id).order_by(
            APIStatisticsHistory.created_at.desc()
        ).all()
        
        if not records:
            return jsonify({"success": False, "error": "No statistics history found"}), 400
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "API Statistics History"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Headers
        headers = [
            'Address',
            'Regrid Available', 'Regrid Unavailable', 'Regrid Total', 'Regrid %',
            'Smarty Available', 'Smarty Unavailable', 'Smarty Total', 'Smarty %',
            'Melissa Available', 'Melissa Unavailable', 'Melissa Total', 'Melissa %'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_num, record in enumerate(records, 2):
            # Calculate percentages (handle division by zero)
            regrid_percentage = int((record.regrid_available / record.regrid_total * 100)) if record.regrid_total > 0 else 0
            smarty_percentage = int((record.smarty_available / record.smarty_total * 100)) if record.smarty_total > 0 else 0
            melissa_percentage = int((record.melissa_available / record.melissa_total * 100)) if record.melissa_total > 0 else 0
            
            data = [
                record.address,
                record.regrid_available,
                record.regrid_unavailable,
                record.regrid_total,
                f"{regrid_percentage}%",
                record.smarty_available,
                record.smarty_unavailable,
                record.smarty_total,
                f"{smarty_percentage}%",
                record.melissa_available,
                record.melissa_unavailable,
                record.melissa_total,
                f"{melissa_percentage}%"
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if col_num <= 2 else "center",
                    vertical="center"
                )
        
        # Column widths
        ws.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"API_Statistics_History_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"[EXCEL EXPORT ERROR] {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500






# =============================================================================
# DATABASE INITIALIZATION
# =============================================================================
def init_database():
    """Initialize database"""
    with app.app_context():
        db.create_all()
        print("[DATABASE] Tables created")
        if User.query.count() == 0:
            test_user = User(
                name='Test User',
                email='test@example.com',
                password_hash=bcrypt.generate_password_hash('Test123!').decode('utf-8')
            )
            db.session.add(test_user)
            db.session.commit()
            print("[DATABASE] Test user: test@example.com / Test123!")


# =============================================================================
# MAIN
# =============================================================================
if __name__ == '__main__':
    init_database()
    TOTAL_FIELDS = len(FIELD_MAPPING)
    print("\n" + "="*80)
    print("PROPERTY DATA TOOL - FIXED VERSION")
    print("="*80)
    print("Server: http://localhost:5000")
    print("Test login: test@example.com / Test123!")
    print("✅ Address validation enabled")
    print(f"✅ ALWAYS shows {TOTAL_FIELDS} total fields")
    print(f"✅ Available + Unavailable = {TOTAL_FIELDS} (VERIFIED)")
    print("="*80 + "\n")
    app.run(debug=True, port=5000)